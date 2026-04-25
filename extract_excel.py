import json
import traceback

data = {}
try:
    import pandas as pd
    xls = pd.ExcelFile('Graphs indicadores.xlsx')
    for sheet in xls.sheet_names:
        df = pd.read_excel('Graphs indicadores.xlsx', sheet_name=sheet, header=None)
        data[sheet] = df.fillna("").values.tolist()
except ImportError:
    try:
        import openpyxl
        wb = openpyxl.load_workbook('Graphs indicadores.xlsx', data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_data = []
            for row in ws.iter_rows(values_only=True):
                sheet_data.append([str(cell) if cell is not None else "" for cell in row])
            data[sheet] = sheet_data
    except Exception as e:
        data["error"] = str(e)
        data["traceback"] = traceback.format_exc()

with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)
